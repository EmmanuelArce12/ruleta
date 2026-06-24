from flask import Flask, render_template, request, redirect, session, Response, jsonify
from werkzeug.security import check_password_hash
from datetime import datetime, date, timedelta
from io import BytesIO
from decimal import Decimal
import os
import re
import threading
import time
import json

import openpyxl
import psycopg2
import psycopg2.extras
from werkzeug.middleware.proxy_fix import ProxyFix

try:
    import qrcode
except ImportError:
    qrcode = None

from dotenv import load_dotenv
from debo import (
    SQLSERVER_DRIVER,
    fetch_ticket_lines_on_date,
    station_debo_ready,
    validate_ticket_invoice_on_date,
)

load_dotenv()

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1)
app.secret_key = os.environ.get('SORTEO_SECRET_KEY') or os.environ.get('FLASK_SECRET_KEY', 'clave_sorteo_electrodomesticos')
DATABASE_URL = os.environ.get('DATABASE_URL')
PUBLIC_BASE_URL = os.environ.get('SORTEO_PUBLIC_BASE_URL', '').rstrip('/')
SORTEO_BASE_PATH = os.environ.get('SORTEO_BASE_PATH', '/sorteo').rstrip('/')

PROMO_COMBUSTIBLES = {'Super', 'Diesel 500', 'Infinia', 'Infinia Diesel'}
INF_INFINIA = {'Infinia', 'Infinia Diesel'}
ACTIVE_STATES = ('PENDIENTE', 'DUDOSO', 'APROBADO', 'DENEGADO')
PROMO_COMBUSTIBLE_KEYS = (
    ('super_litros', 'Super', 'nafta'),
    ('infinia_litros', 'Infinia', 'nafta'),
    ('diesel_500_litros', 'Diesel 500', 'diesel'),
    ('infinia_diesel_litros', 'Infinia Diesel', 'diesel'),
)

ALLOWED_BRANCHES_BY_STATION = {
    'echeverria': {'17', '18', '19'},
    'satragno': {'16', '17', '18', '25', '26'},
}


def get_db():
    return psycopg2.connect(DATABASE_URL)


def get_sorteo_config(estacion_id):
    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute('SELECT * FROM sorteo_config WHERE estacion_id = %s', (estacion_id,))
    row = c.fetchone()
    conn.close()
    return row


def get_station(estacion_id):
    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute('SELECT * FROM estaciones WHERE id = %s', (estacion_id,))
    row = c.fetchone()
    conn.close()
    return row


def get_allowed_invoice_branches(station: dict | None) -> list[str]:
    name = ((station or {}).get('nombre') or '').strip().lower()
    for key, branches in ALLOWED_BRANCHES_BY_STATION.items():
        if key in name:
            return sorted(branches, key=lambda value: int(value))
    return []


def init_db():
    conn = get_db()
    conn.autocommit = True
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS sorteo_config (
            estacion_id INTEGER PRIMARY KEY REFERENCES estaciones(id) ON DELETE CASCADE,
            minimo_litros NUMERIC(10, 3) DEFAULT 0,
            intervalo_horas INTEGER DEFAULT 4,
            consulta_automatica_hora TIME DEFAULT TIME '04:00',
            promocion_desde DATE,
            promocion_hasta DATE,
            activo BOOLEAN DEFAULT TRUE,
            detenido BOOLEAN DEFAULT FALSE,
            ultima_consulta TIMESTAMP,
            consulta_en_curso BOOLEAN DEFAULT FALSE,
            estado_consulta TEXT,
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            actualizado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute("ALTER TABLE sorteo_config ADD COLUMN IF NOT EXISTS promocion_desde DATE")
    c.execute("ALTER TABLE sorteo_config ADD COLUMN IF NOT EXISTS promocion_hasta DATE")
    c.execute("ALTER TABLE sorteo_config ADD COLUMN IF NOT EXISTS consulta_automatica_hora TIME DEFAULT TIME '04:00'")
    c.execute("ALTER TABLE sorteo_config ADD COLUMN IF NOT EXISTS consulta_en_curso BOOLEAN DEFAULT FALSE")
    c.execute("ALTER TABLE sorteo_config ADD COLUMN IF NOT EXISTS estado_consulta TEXT")
    c.execute("""
        UPDATE sorteo_config
        SET consulta_en_curso = FALSE,
            estado_consulta = CASE
                WHEN consulta_en_curso = TRUE THEN 'La consulta anterior se interrumpio por reinicio del servicio.'
                ELSE estado_consulta
            END,
            actualizado_en = CURRENT_TIMESTAMP
        WHERE consulta_en_curso = TRUE
    """)
    c.execute('''
        CREATE TABLE IF NOT EXISTS sorteo_participantes (
            id SERIAL PRIMARY KEY,
            estacion_id INTEGER REFERENCES estaciones(id) ON DELETE CASCADE,
            ticket_fecha DATE NOT NULL,
            ticket_hora TIMESTAMP,
            numero_factura TEXT NOT NULL,
            nombre TEXT NOT NULL,
            apellido TEXT NOT NULL,
            dni TEXT,
            telefono TEXT NOT NULL,
            email TEXT NOT NULL,
            acepta_promociones BOOLEAN DEFAULT FALSE,
            estado TEXT DEFAULT 'PENDIENTE',
            combustible TEXT,
            litros NUMERIC(10, 3),
            pago_app_ypf BOOLEAN,
            pago_electronico BOOLEAN,
            medio_pago TEXT,
            payment_type TEXT,
            vendedor TEXT,
            factura_real TEXT,
            chances INTEGER DEFAULT 1,
            tipo_comprobante TEXT,
            letra_fiscal TEXT,
            device_token TEXT,
            ip_registro TEXT,
            user_agent TEXT,
            sospecha_dispositivo BOOLEAN DEFAULT FALSE,
            detalle_validacion TEXT,
            consulta_at TIMESTAMP,
            conteo_id INTEGER,
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS ticket_fecha DATE")
    c.execute("UPDATE sorteo_participantes SET ticket_fecha = COALESCE(ticket_fecha, ticket_hora::date) WHERE ticket_fecha IS NULL")
    c.execute("ALTER TABLE sorteo_participantes ALTER COLUMN ticket_fecha SET NOT NULL")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS ticket_hora TIMESTAMP")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS acepta_promociones BOOLEAN DEFAULT FALSE")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS pago_electronico BOOLEAN")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS medio_pago TEXT")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS payment_type TEXT")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS vendedor TEXT")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS factura_real TEXT")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS chances INTEGER DEFAULT 1")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS tipo_comprobante TEXT")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS letra_fiscal TEXT")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS device_token TEXT")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS ip_registro TEXT")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS user_agent TEXT")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS sospecha_dispositivo BOOLEAN DEFAULT FALSE")
    c.execute("ALTER TABLE sorteo_participantes ADD COLUMN IF NOT EXISTS composicion_ticket TEXT")
    c.execute('''
        CREATE TABLE IF NOT EXISTS sorteo_conteos (
            id SERIAL PRIMARY KEY,
            estacion_id INTEGER REFERENCES estaciones(id) ON DELETE CASCADE,
            iniciado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            promocion_desde DATE,
            promocion_hasta DATE,
            total INTEGER DEFAULT 0,
            aprobados INTEGER DEFAULT 0,
            denegados INTEGER DEFAULT 0,
            pendientes INTEGER DEFAULT 0,
            dudosos INTEGER DEFAULT 0,
            snapshot_json TEXT,
            ranking_json TEXT
        )
    ''')
    c.execute("ALTER TABLE sorteo_conteos ADD COLUMN IF NOT EXISTS dudosos INTEGER DEFAULT 0")
    c.execute("ALTER TABLE sorteo_conteos ADD COLUMN IF NOT EXISTS promocion_desde DATE")
    c.execute("ALTER TABLE sorteo_conteos ADD COLUMN IF NOT EXISTS promocion_hasta DATE")
    c.execute("ALTER TABLE sorteo_conteos ADD COLUMN IF NOT EXISTS ranking_json TEXT")
    c.execute('''
        CREATE INDEX IF NOT EXISTS idx_sorteo_participantes_estacion_activo
        ON sorteo_participantes(estacion_id, conteo_id, creado_en DESC)
    ''')
    c.execute('DROP INDEX IF EXISTS idx_sorteo_participantes_factura_estacion')
    c.execute('DROP INDEX IF EXISTS idx_sorteo_participantes_factura_estacion_activa')
    c.execute('DROP INDEX IF EXISTS idx_sorteo_participantes_ticket_factura_activa')
    c.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_sorteo_participantes_fecha_factura_activa
        ON sorteo_participantes(estacion_id, ticket_fecha, numero_factura)
        WHERE conteo_id IS NULL
    ''')
    conn.close()


def ensure_config(estacion_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('INSERT INTO sorteo_config (estacion_id) VALUES (%s) ON CONFLICT (estacion_id) DO NOTHING', (estacion_id,))
    c.execute('''
        UPDATE sorteo_config
        SET promocion_desde = COALESCE(promocion_desde, CURRENT_DATE),
            promocion_hasta = COALESCE(promocion_hasta, CURRENT_DATE + INTERVAL '30 days'),
            consulta_automatica_hora = COALESCE(consulta_automatica_hora, TIME '04:00'),
            actualizado_en = CURRENT_TIMESTAMP
        WHERE estacion_id = %s
    ''', (estacion_id,))
    conn.commit()
    conn.close()


def admin_required():
    return 'sorteo_estacion_id' in session


def sorteo_path(path=''):
    if not path.startswith('/'):
        path = '/' + path
    return f'{SORTEO_BASE_PATH}{path}' if SORTEO_BASE_PATH else path


def public_url(estacion_id):
    path = sorteo_path(f'/{estacion_id}')
    if PUBLIC_BASE_URL:
        return f'{PUBLIC_BASE_URL}{path}'
    forwarded_proto = (request.headers.get('X-Forwarded-Proto') or request.scheme or 'https').split(',')[0].strip()
    forwarded_host = (request.headers.get('X-Forwarded-Host') or request.headers.get('Host') or request.host).split(',')[0].strip()
    return f'{forwarded_proto}://{forwarded_host}{path}'


def qr_data_url(text):
    if qrcode is None:
        return None
    img = qrcode.make(text)
    salida = BytesIO()
    img.save(salida, format='PNG')
    import base64
    return 'data:image/png;base64,' + base64.b64encode(salida.getvalue()).decode('ascii')


def parse_ticket_date(fecha):
    return datetime.strptime(fecha, '%Y-%m-%d').date()


def normalize_invoice_branch_digits(raw_value):
    digits = re.sub(r'\D+', '', raw_value or '')
    digits = digits.lstrip('0') or '0'
    if len(digits) > 2:
        digits = digits.rstrip('0') or digits
    digits = digits.lstrip('0') or '0'
    return str(int(digits))


def normalize_invoice_serial_digits(raw_value):
    digits = re.sub(r'\D+', '', raw_value or '')
    digits = digits.lstrip('0') or '0'
    return str(int(digits))


def build_canonical_invoice(branch_raw, serial_raw):
    sucursal = int(normalize_invoice_branch_digits(branch_raw))
    numero = int(normalize_invoice_serial_digits(serial_raw))
    return {
        'sucursal': sucursal,
        'numero': numero,
        'canonical': f'{sucursal}-{numero}',
    }


def normalize_invoice_number(raw_value):
    raw_value = (raw_value or '').strip()
    groups = re.findall(r'\d+', raw_value)
    if not groups:
        raise ValueError('Ingresa el numero de factura con punto de venta y numero.')
    if len(groups) >= 2:
        return build_canonical_invoice(groups[-2], groups[-1])
    else:
        digits = groups[0]
        if len(digits) <= 6:
            raise ValueError('El numero de factura debe incluir punto de venta y numero.')
        return build_canonical_invoice(digits[:-6], digits[-6:])


def parse_invoice_from_form(form):
    punto_venta = form.get('numero_factura_sucursal')
    numero = form.get('numero_factura_numero')
    if punto_venta is not None or numero is not None:
        if not (punto_venta or '').strip() or not (numero or '').strip():
            raise ValueError('Completa punto de venta y numero de factura.')
        return build_canonical_invoice(punto_venta, numero)
    return normalize_invoice_number(form.get('numero_factura'))


def decimal_to_float(value):
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    return float(value or 0)


def yes_no_blank(value):
    if value is True:
        return 'Si'
    if value is False:
        return 'No'
    return ''


def split_timestamp(value):
    if not value:
        return '', ''
    text = str(value)[:19]
    parts = text.split(' ')
    if len(parts) == 2:
        return parts[0], parts[1]
    return text, ''


def compute_chances(combustible, pago_app_ypf):
    base = 2 if combustible in INF_INFINIA else 1
    return min(3, base + (1 if pago_app_ypf else 0))


def build_ticket_composition(factura):
    items = []
    family_totals = {'nafta': 0.0, 'diesel': 0.0}
    for key, label, family in PROMO_COMBUSTIBLE_KEYS:
        liters = decimal_to_float(factura.get(key))
        if liters > 0:
            items.append({'combustible': label, 'litros': round(liters, 4), 'familia': family})
            family_totals[family] += liters
    return {'items': items, 'family_totals': family_totals}


def serialize_ticket_composition(factura):
    return json.dumps(build_ticket_composition(factura), ensure_ascii=False)


def mixed_ticket_detail(factura):
    composition = build_ticket_composition(factura)
    parts = [f"{item['combustible']}: {item['litros']:.4f} lt" for item in composition['items']]
    return ' | '.join(parts)


def technical_validation_error(message):
    return message.startswith('Falta configurar') or message.startswith('Error consultando DEBO:')


def suspicious_device_message():
    return 'Registro marcado como dudoso: este dispositivo ya supera el limite permitido de cargas en la promocion.'


def get_client_ip():
    forwarded = request.headers.get('X-Forwarded-For', '')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.remote_addr or ''


def get_active_promo_bounds(config):
    if config and config.get('promocion_desde') and config.get('promocion_hasta'):
        return config['promocion_desde'], config['promocion_hasta']
    today = date.today()
    return today, today


def count_device_registrations(cursor, estacion_id, config, device_token, exclude_id=None):
    if not device_token:
        return 0
    promo_desde, promo_hasta = get_active_promo_bounds(config)
    sql = '''
        SELECT COUNT(*)
        FROM sorteo_participantes
        WHERE estacion_id = %s
          AND conteo_id IS NULL
          AND device_token = %s
          AND ticket_fecha >= %s
          AND ticket_fecha <= %s
    '''
    params = [estacion_id, device_token, promo_desde, promo_hasta]
    if exclude_id is not None:
        sql += ' AND id <> %s'
        params.append(exclude_id)
    cursor.execute(sql, params)
    row = cursor.fetchone()
    if isinstance(row, dict):
        return int(row.get('count', 0))
    return int(row[0]) if row else 0


def is_suspicious_device(cursor, participante, config):
    return count_device_registrations(
        cursor,
        participante['estacion_id'],
        config,
        participante.get('device_token'),
        exclude_id=participante['id'],
    ) >= 2


def consultar_facturacion(estacion_id, ticket_fecha, numero_factura):
    station = get_station(estacion_id)
    if not station_debo_ready(station):
        return None, 'Falta configurar IP, base, usuario o clave de DEBO para esta estacion.'
    try:
        factura = normalize_invoice_number(numero_factura)
        validate_allowed_invoice_branch(get_station(estacion_id), factura)
    except ValueError as exc:
        return None, str(exc)
    try:
        return validate_ticket_invoice_on_date(
            station,
            datetime.combine(ticket_fecha, datetime.min.time()),
            factura['sucursal'],
            factura['numero'],
            include_remitos=bool(station.get('debo_allow_remitos')),
        )
    except Exception as exc:
        return None, f'Error consultando DEBO: {exc}'


def consultar_detalle_ticket(estacion_id, ticket_fecha, numero_factura):
    station = get_station(estacion_id)
    if not station_debo_ready(station):
        return None, 'Falta configurar IP, base, usuario o clave de DEBO para esta estacion.'
    try:
        factura = normalize_invoice_number(numero_factura)
        validate_allowed_invoice_branch(get_station(estacion_id), factura)
    except ValueError as exc:
        return None, str(exc)
    try:
        rows = fetch_ticket_lines_on_date(
            station,
            datetime.combine(ticket_fecha, datetime.min.time()),
            factura['sucursal'],
            factura['numero'],
            include_remitos=bool(station.get('debo_allow_remitos')),
        )
    except Exception as exc:
        return None, f'Error consultando DEBO: {exc}'
    if not rows:
        return None, 'No se encontro el detalle del ticket para esa fecha y numero.'
    return rows, None


def classify_ticket_match(factura, minimo_litros):
    promo_lineas = int(factura.get('promo_lineas') or 0)
    minimum = float(minimo_litros or 0)
    composition = build_ticket_composition(factura)
    items = composition['items']
    family_totals = composition['family_totals']
    total_promo_liters = sum(item['litros'] for item in items)
    pago_app_ypf = bool(factura.get('pago_app_ypf'))

    if promo_lineas <= 0 or not items:
        return {
            'estado': 'DENEGADO',
            'detalle': 'La factura no corresponde a un combustible participante.',
            'combustible': factura.get('combustible'),
            'litros': decimal_to_float(factura.get('litros')),
            'chances': 0,
            'composicion_ticket': serialize_ticket_composition(factura),
        }

    qualifiers = [family for family, liters in family_totals.items() if liters >= minimum]
    non_zero_families = [family for family, liters in family_totals.items() if liters > 0]

    if not qualifiers:
        if len(non_zero_families) > 1:
            return {
                'estado': 'DENEGADO',
                'detalle': f"Mezcla nafta y diesel no permitida. {mixed_ticket_detail(factura)}.",
                'combustible': 'Mixto',
                'litros': total_promo_liters,
                'chances': 0,
                'composicion_ticket': serialize_ticket_composition(factura),
            }
        return {
            'estado': 'DENEGADO',
            'detalle': f'Litros insuficientes: {total_promo_liters:.3f} de minimo {minimum:.3f}.',
            'combustible': factura.get('combustible'),
            'litros': total_promo_liters,
            'chances': 0,
            'composicion_ticket': serialize_ticket_composition(factura),
        }

    if len(qualifiers) > 1:
        return {
            'estado': 'DENEGADO',
            'detalle': f"Mezcla nafta y diesel no permitida. {mixed_ticket_detail(factura)}.",
            'combustible': 'Mixto',
            'litros': total_promo_liters,
            'chances': 0,
            'composicion_ticket': serialize_ticket_composition(factura),
        }

    qualifying_family = qualifiers[0]
    family_items = [item for item in items if item['familia'] == qualifying_family]
    family_labels = {item['combustible'] for item in family_items}
    family_liters = sum(item['litros'] for item in family_items)

    if qualifying_family == 'nafta':
        if family_labels == {'Infinia'}:
            combustible = 'Infinia'
            base_chances = 2
        elif family_labels == {'Super'}:
            combustible = 'Super'
            base_chances = 1
        else:
            combustible = 'Nafta mixta'
            base_chances = 1
    else:
        if family_labels == {'Infinia Diesel'}:
            combustible = 'Infinia Diesel'
            base_chances = 2
        elif family_labels == {'Diesel 500'}:
            combustible = 'Diesel 500'
            base_chances = 1
        else:
            combustible = 'Diesel mixto'
            base_chances = 1

    chances = min(3, base_chances + (1 if pago_app_ypf else 0))
    detalle = 'Validado correctamente por combustible participante.'
    if len(non_zero_families) > 1:
        detalle = f"Validado por {combustible}. La otra familia no suma por mezcla. {mixed_ticket_detail(factura)}."
    elif len(family_labels) > 1:
        detalle = f"Validado por suma de la misma familia. {mixed_ticket_detail(factura)}."

    return {
        'estado': 'APROBADO',
        'detalle': detalle,
        'combustible': combustible,
        'litros': family_liters,
        'chances': chances,
        'composicion_ticket': serialize_ticket_composition(factura),
    }


def query_participantes(cursor, eid, archived=False):
    if archived:
        cursor.execute('SELECT * FROM sorteo_participantes WHERE estacion_id = %s ORDER BY creado_en DESC', (eid,))
    else:
        cursor.execute('SELECT * FROM sorteo_participantes WHERE estacion_id = %s AND conteo_id IS NULL ORDER BY creado_en DESC', (eid,))
    return cursor.fetchall()


def build_seller_ranking(rows):
    ranking = {}
    for row in rows:
        vendedor = (row.get('vendedor') or 'Sin vendedor').strip() or 'Sin vendedor'
        ranking[vendedor] = ranking.get(vendedor, 0) + 1
    return [
        {'vendedor': vendedor, 'facturas_aprobadas': total}
        for vendedor, total in sorted(ranking.items(), key=lambda item: (-item[1], item[0]))
    ]


def format_ticket_detail_rows(rows):
    formatted = []
    for row in rows or []:
        quantity = decimal_to_float(row.get('cantidad'))
        formatted.append({
            'sector': row.get('sector'),
            'articulo': row.get('articulo'),
            'producto': row.get('producto') or 'Sin descripcion',
            'cantidad': round(quantity, 4),
            'es_combustible_participante': bool(row.get('es_combustible_participante')),
        })
    return formatted


EXPORT_HEADERS = [
    'Registro',
    'Estado',
    'Fecha ticket',
    'Factura cliente',
    'Factura DEBO',
    'Tipo',
    'Letra',
    'Vendedor',
    'Combustible',
    'Litros',
    'Medio pago',
    'App YPF',
    'Chances',
    'Device token',
    'IP',
    'User agent',
    'Sospecha dispositivo',
    'Nombre',
    'Apellido',
    'DNI',
    'Telefono',
    'Email',
    'Acepta promociones',
    'Detalle',
    'Consultado',
    'Conteo',
    'Repeticion exportada',
]


def build_export_values(row, repetition=''):
    return [
        str(row.get('creado_en') or '')[:19],
        row.get('estado') or '',
        str(row.get('ticket_fecha') or ''),
        row.get('numero_factura') or '',
        row.get('factura_real') or '',
        row.get('tipo_comprobante') or '',
        row.get('letra_fiscal') or '',
        row.get('vendedor') or '',
        row.get('combustible') or '',
        row.get('litros'),
        row.get('medio_pago') or '',
        yes_no_blank(row.get('pago_app_ypf')),
        row.get('chances') or '',
        row.get('device_token') or '',
        row.get('ip_registro') or '',
        row.get('user_agent') or '',
        'Si' if row.get('sospecha_dispositivo') else 'No',
        row.get('nombre') or '',
        row.get('apellido') or '',
        row.get('dni') or '',
        row.get('telefono') or '',
        row.get('email') or '',
        'Si' if row.get('acepta_promociones') else 'No',
        row.get('detalle_validacion') or '',
        str(row.get('consulta_at') or '')[:19] if row.get('consulta_at') else '',
        row.get('conteo_id'),
        repetition,
    ]


def append_export_sheet(workbook, title, rows, ponderado=False):
    ws = workbook.create_sheet(title=title)
    ws.append(EXPORT_HEADERS)
    for row in rows:
        repeats = max(int(row.get('chances') or 1), 1) if ponderado and (row.get('estado') == 'APROBADO') else 1
        for idx in range(repeats):
            repetition = idx + 1 if ponderado and (row.get('estado') == 'APROBADO') else ''
            ws.append(build_export_values(row, repetition=repetition))
    return ws


def append_archive_summary_sheet(workbook, conteo):
    ws = workbook.create_sheet(title='Resumen')
    ws.append(['Campo', 'Valor'])
    ws.append(['Archivado ID', conteo.get('id')])
    ws.append(['Fecha archivado', str(conteo.get('iniciado_en') or '')[:19]])
    ws.append(['Promocion desde', str(conteo.get('promocion_desde') or '')])
    ws.append(['Promocion hasta', str(conteo.get('promocion_hasta') or '')])
    ws.append(['Total', conteo.get('total') or 0])
    ws.append(['Aprobados', conteo.get('aprobados') or 0])
    ws.append(['Denegados', conteo.get('denegados') or 0])
    ws.append(['Pendientes', conteo.get('pendientes') or 0])
    ws.append(['Dudosos', conteo.get('dudosos') or 0])
    return ws


def append_archive_ranking_sheet(workbook, ranking_rows):
    ws = workbook.create_sheet(title='Ranking')
    ws.append(['Puesto', 'Vendedor', 'Facturas aprobadas'])
    for index, row in enumerate(ranking_rows or [], start=1):
        ws.append([index, row.get('vendedor') or 'Sin vendedor', row.get('facturas_aprobadas') or 0])
    return ws


def build_admin_context(eid):
    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute('SELECT * FROM sorteo_config WHERE estacion_id = %s', (eid,))
    config = c.fetchone()
    c.execute('SELECT * FROM estaciones WHERE id = %s', (eid,))
    station = c.fetchone()
    participantes = query_participantes(c, eid, archived=False)
    c.execute('SELECT * FROM sorteo_conteos WHERE estacion_id = %s ORDER BY iniciado_en DESC LIMIT 20', (eid,))
    conteos = c.fetchall()
    conn.close()

    aprobados = [p for p in participantes if p['estado'] == 'APROBADO']
    denegados = [p for p in participantes if p['estado'] == 'DENEGADO']
    pendientes = [p for p in participantes if p['estado'] == 'PENDIENTE']
    dudosos = [p for p in participantes if p['estado'] == 'DUDOSO']
    url = public_url(eid)

    return {
        'nombre_estacion': session['sorteo_estacion_nombre'],
        'admin_notice': session.pop('sorteo_admin_notice', None),
        'config': config,
        'estacion': station,
        'aprobados': aprobados,
        'denegados': denegados,
        'pendientes': pendientes,
        'dudosos': dudosos,
        'conteos': conteos,
        'public_url': url,
        'qr': qr_data_url(url),
        'facturacion_configurada': station_debo_ready(station),
        'base_path': SORTEO_BASE_PATH,
        'sqlserver_driver': SQLSERVER_DRIVER,
        'seller_ranking': build_seller_ranking(aprobados),
        'totales': {
            'aprobados': len(aprobados),
            'denegados': len(denegados),
            'pendientes': len(pendientes),
            'dudosos': len(dudosos),
        },
        'split_timestamp': split_timestamp,
        'yes_no_blank': yes_no_blank,
    }


def validate_allowed_invoice_branch(station: dict | None, factura: dict):
    allowed = get_allowed_invoice_branches(station)
    if allowed and str(factura['sucursal']) not in allowed:
        raise ValueError('El punto de venta no corresponde a esta estacion.')


def reset_participante_validation(cursor, participante_id, estado='PENDIENTE', detalle='Esperando validacion.'):
    cursor.execute('''
        UPDATE sorteo_participantes
        SET estado = %s,
            combustible = NULL,
            litros = NULL,
            pago_app_ypf = NULL,
            pago_electronico = NULL,
            medio_pago = NULL,
            payment_type = NULL,
            vendedor = NULL,
            factura_real = NULL,
            chances = 0,
            tipo_comprobante = NULL,
            letra_fiscal = NULL,
            composicion_ticket = NULL,
            detalle_validacion = %s,
            consulta_at = CURRENT_TIMESTAMP
        WHERE id = %s
    ''', (estado, detalle, participante_id))


def requeue_participante_with_invoice(cursor, estacion_id, participante_id, numero_factura, ticket_fecha=None):
    factura = normalize_invoice_number(numero_factura)
    cursor.execute('''
        SELECT id, ticket_fecha
        FROM sorteo_participantes
        WHERE id = %s
          AND estacion_id = %s
          AND conteo_id IS NULL
          AND estado IN ('DENEGADO', 'DUDOSO')
    ''', (participante_id, estacion_id))
    row = cursor.fetchone()
    if not row:
        raise ValueError('El registro ya no esta disponible para reproceso.')

    row_id = row['id'] if isinstance(row, dict) else row[0]
    current_ticket_fecha = row['ticket_fecha'] if isinstance(row, dict) else row[1]
    target_ticket_fecha = parse_ticket_date(ticket_fecha) if ticket_fecha else current_ticket_fecha
    cursor.execute('''
        SELECT id
        FROM sorteo_participantes
        WHERE estacion_id = %s
          AND conteo_id IS NULL
          AND ticket_fecha = %s
          AND numero_factura = %s
          AND id <> %s
        LIMIT 1
    ''', (estacion_id, target_ticket_fecha, factura['canonical'], row_id))
    if cursor.fetchone():
        raise ValueError('Ya existe otra carga activa con esa fecha y factura.')

    cursor.execute('''
        UPDATE sorteo_participantes
        SET numero_factura = %s,
            ticket_fecha = %s,
            ticket_hora = %s
        WHERE id = %s
    ''', (factura['canonical'], target_ticket_fecha, datetime.combine(target_ticket_fecha, datetime.min.time()), row_id))
    reset_participante_validation(cursor, row_id)
    return factura['canonical'], str(target_ticket_fecha)


def validar_pendientes(estacion_id=None):
    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if estacion_id:
        c.execute('''
            SELECT p.*, cfg.minimo_litros, cfg.promocion_desde, cfg.promocion_hasta
            FROM sorteo_participantes p
            JOIN sorteo_config cfg ON cfg.estacion_id = p.estacion_id
            WHERE p.estado = 'PENDIENTE' AND p.conteo_id IS NULL AND p.estacion_id = %s
            ORDER BY p.creado_en ASC
        ''', (estacion_id,))
    else:
        c.execute('''
            SELECT p.*, cfg.minimo_litros, cfg.promocion_desde, cfg.promocion_hasta
            FROM sorteo_participantes p
            JOIN sorteo_config cfg ON cfg.estacion_id = p.estacion_id
            WHERE p.estado = 'PENDIENTE' AND p.conteo_id IS NULL
            ORDER BY p.creado_en ASC
        ''')
    pendientes = c.fetchall()

    for participante in pendientes:
        promo_config = {
            'promocion_desde': participante.get('promocion_desde'),
            'promocion_hasta': participante.get('promocion_hasta'),
        }
        factura, error = consultar_facturacion(
            participante['estacion_id'],
            participante['ticket_fecha'],
            participante['numero_factura'],
        )
        if error:
            nuevo_estado = 'PENDIENTE' if technical_validation_error(error) else 'DENEGADO'
            reset_participante_validation(c, participante['id'], estado=nuevo_estado, detalle=error)
            continue

        evaluation = classify_ticket_match(factura, participante['minimo_litros'])
        estado = evaluation['estado']
        detalle = evaluation['detalle']
        combustible = evaluation['combustible']
        pago_app_ypf = bool(factura.get('pago_app_ypf'))
        pago_electronico = bool(factura.get('pago_electronico'))
        medio_pago = factura.get('medio_pago') or ('App YPF' if pago_app_ypf else 'Contado / no electronico')
        chances = evaluation['chances'] if estado == 'APROBADO' else 0
        sospecha_dispositivo = is_suspicious_device(c, participante, promo_config)

        if estado == 'APROBADO' and sospecha_dispositivo:
            estado = 'DUDOSO'
            detalle = suspicious_device_message()

        c.execute('''
            UPDATE sorteo_participantes
            SET estado = %s,
                combustible = %s,
                litros = %s,
                pago_app_ypf = %s,
                pago_electronico = %s,
                medio_pago = %s,
                payment_type = %s,
                vendedor = %s,
                factura_real = %s,
                chances = %s,
                tipo_comprobante = %s,
                letra_fiscal = %s,
                composicion_ticket = %s,
                sospecha_dispositivo = %s,
                detalle_validacion = %s,
                consulta_at = CURRENT_TIMESTAMP
            WHERE id = %s
        ''', (
            estado,
            combustible,
            evaluation['litros'],
            pago_app_ypf,
            pago_electronico,
            medio_pago,
            factura.get('payment_type'),
            factura.get('vendedor'),
            factura.get('numero_factura'),
            chances,
            factura.get('tipo_comprobante'),
            factura.get('letra_fiscal'),
            evaluation.get('composicion_ticket'),
            sospecha_dispositivo,
            detalle,
            participante['id'],
        ))

    if estacion_id:
        c.execute('UPDATE sorteo_config SET ultima_consulta = CURRENT_TIMESTAMP, actualizado_en = CURRENT_TIMESTAMP WHERE estacion_id = %s', (estacion_id,))
    else:
        c.execute('UPDATE sorteo_config SET ultima_consulta = CURRENT_TIMESTAMP, actualizado_en = CURRENT_TIMESTAMP')
    conn.commit()
    conn.close()
    return len(pendientes)


def set_consulta_status(estacion_id, en_curso, mensaje):
    conn = get_db()
    c = conn.cursor()
    c.execute('''
        UPDATE sorteo_config
        SET consulta_en_curso = %s,
            estado_consulta = %s,
            actualizado_en = CURRENT_TIMESTAMP
        WHERE estacion_id = %s
    ''', (en_curso, mensaje, estacion_id))
    conn.commit()
    conn.close()


def ejecutar_consulta_manual_async(estacion_id):
    try:
        procesados = validar_pendientes(estacion_id)
        set_consulta_status(estacion_id, False, f'Consulta manual finalizada. Pendientes revisados: {procesados}.')
    except Exception as exc:
        set_consulta_status(estacion_id, False, f'Error en consulta manual: {exc}')


def iniciar_consulta_en_segundo_plano(estacion_id, origen='manual'):
    config = get_sorteo_config(estacion_id) or {}
    if config.get('consulta_en_curso'):
        updated_at = config.get('actualizado_en')
        if updated_at and isinstance(updated_at, datetime) and updated_at >= datetime.now(updated_at.tzinfo) - timedelta(minutes=20):
            return False, 'Ya hay una consulta en curso para esta estacion.'
        set_consulta_status(estacion_id, False, 'Se libero una consulta previa que habia quedado colgada.')

    if origen == 'manual':
        status_text = 'Consulta manual en curso. La pagina puede recargarse mientras se procesa.'
        notice_text = 'Consulta manual iniciada en segundo plano.'
    else:
        status_text = 'Consulta automatica en curso.'
        notice_text = 'Consulta automatica iniciada.'

    set_consulta_status(estacion_id, True, status_text)
    worker = threading.Thread(target=ejecutar_consulta_manual_async, args=(estacion_id,), daemon=True)
    worker.start()
    return True, notice_text


def scheduler_loop():
    while True:
        time.sleep(60)
        try:
            conn = get_db()
            c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            c.execute('''
                SELECT estacion_id
                FROM sorteo_config
                WHERE activo = TRUE
                  AND detenido = FALSE
                  AND CURRENT_TIMESTAMP >= date_trunc('day', CURRENT_TIMESTAMP) + COALESCE(consulta_automatica_hora, TIME '04:00')
                  AND (
                    ultima_consulta IS NULL OR
                    ultima_consulta < date_trunc('day', CURRENT_TIMESTAMP) + COALESCE(consulta_automatica_hora, TIME '04:00')
                  )
            ''')
            estaciones = c.fetchall()
            conn.close()
            for est in estaciones:
                validar_pendientes(est['estacion_id'])
        except Exception as exc:
            print('Error en scheduler sorteo:', exc)


@app.route('/')
@app.route('/sorteo')
def root():
    return redirect(sorteo_path('/admin'))


@app.route('/admin/login', methods=['GET', 'POST'])
@app.route('/sorteo/admin/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        usuario = request.form['usuario'].lower().strip()
        password = request.form['password']
        conn = get_db()
        c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        c.execute('SELECT * FROM estaciones WHERE admin_user = %s', (usuario,))
        estacion = c.fetchone()
        conn.close()
        if estacion and check_password_hash(estacion['admin_pass'], password):
            session['sorteo_estacion_id'] = estacion['id']
            session['sorteo_estacion_nombre'] = estacion['nombre']
            ensure_config(estacion['id'])
            return redirect(sorteo_path('/admin'))
        error = 'Credenciales incorrectas.'
    return render_template('sorteo_login.html', error=error)


@app.route('/admin/logout')
@app.route('/sorteo/admin/logout')
def logout():
    session.clear()
    return redirect(sorteo_path('/admin/login'))


@app.route('/admin')
@app.route('/sorteo/admin')
def admin():
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    eid = session['sorteo_estacion_id']
    ensure_config(eid)
    return render_template('sorteo_admin.html', **build_admin_context(eid))


@app.route('/admin/config', methods=['POST'])
@app.route('/sorteo/admin/config', methods=['POST'])
def configurar():
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    eid = session['sorteo_estacion_id']
    minimo_litros = request.form.get('minimo_litros') or 0
    consulta_automatica_hora = request.form.get('consulta_automatica_hora') or '04:00'
    promocion_desde = request.form.get('promocion_desde') or None
    promocion_hasta = request.form.get('promocion_hasta') or None
    conn = get_db()
    c = conn.cursor()
    c.execute('''
        UPDATE sorteo_config
        SET minimo_litros = %s,
            consulta_automatica_hora = %s,
            promocion_desde = %s,
            promocion_hasta = %s,
            actualizado_en = CURRENT_TIMESTAMP
        WHERE estacion_id = %s
    ''', (minimo_litros, consulta_automatica_hora, promocion_desde, promocion_hasta, eid))
    conn.commit()
    conn.close()
    return redirect(sorteo_path('/admin'))


@app.route('/admin/forzar-consulta', methods=['POST'])
@app.route('/sorteo/admin/forzar-consulta', methods=['POST'])
def forzar_consulta():
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    estacion_id = session['sorteo_estacion_id']
    _, notice = iniciar_consulta_en_segundo_plano(estacion_id, origen='manual')
    session['sorteo_admin_notice'] = notice
    return redirect(sorteo_path('/admin'))


@app.route('/admin/detener', methods=['POST'])
@app.route('/sorteo/admin/detener', methods=['POST'])
def detener():
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE sorteo_config SET detenido = TRUE, actualizado_en = CURRENT_TIMESTAMP WHERE estacion_id = %s', (session['sorteo_estacion_id'],))
    conn.commit()
    conn.close()
    return redirect(sorteo_path('/admin'))


@app.route('/admin/reanudar', methods=['POST'])
@app.route('/sorteo/admin/reanudar', methods=['POST'])
def reanudar():
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE sorteo_config SET detenido = FALSE, actualizado_en = CURRENT_TIMESTAMP WHERE estacion_id = %s', (session['sorteo_estacion_id'],))
    conn.commit()
    conn.close()
    return redirect(sorteo_path('/admin'))


def archive_current_promo(cursor, eid):
    cursor.execute('SELECT promocion_desde, promocion_hasta FROM sorteo_config WHERE estacion_id = %s', (eid,))
    config = cursor.fetchone() or {}
    rows = query_participantes(cursor, eid, archived=False)
    ranking_rows = build_seller_ranking([row for row in rows if row['estado'] == 'APROBADO'])
    total = len(rows)
    aprobados = sum(1 for row in rows if row['estado'] == 'APROBADO')
    denegados = sum(1 for row in rows if row['estado'] == 'DENEGADO')
    pendientes = sum(1 for row in rows if row['estado'] == 'PENDIENTE')
    dudosos = sum(1 for row in rows if row['estado'] == 'DUDOSO')
    snapshot = json.dumps([dict(row) for row in rows], default=str, ensure_ascii=False)
    ranking_json = json.dumps(ranking_rows, default=str, ensure_ascii=False)
    cursor.execute('''
        INSERT INTO sorteo_conteos (estacion_id, promocion_desde, promocion_hasta, total, aprobados, denegados, pendientes, dudosos, snapshot_json, ranking_json)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
    ''', (
        eid,
        config.get('promocion_desde'),
        config.get('promocion_hasta'),
        total,
        aprobados,
        denegados,
        pendientes,
        dudosos,
        snapshot,
        ranking_json,
    ))
    conteo_id = cursor.fetchone()['id']
    cursor.execute('UPDATE sorteo_participantes SET conteo_id = %s WHERE estacion_id = %s AND conteo_id IS NULL', (conteo_id, eid))


@app.route('/admin/nueva-promocion', methods=['POST'])
@app.route('/sorteo/admin/nueva-promocion', methods=['POST'])
def nueva_promocion():
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    eid = session['sorteo_estacion_id']
    promocion_desde = request.form.get('promocion_desde') or None
    promocion_hasta = request.form.get('promocion_hasta') or None
    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    archive_current_promo(c, eid)
    c.execute('''
        UPDATE sorteo_config
        SET promocion_desde = %s,
            promocion_hasta = %s,
            actualizado_en = CURRENT_TIMESTAMP
        WHERE estacion_id = %s
    ''', (promocion_desde, promocion_hasta, eid))
    conn.commit()
    conn.close()
    return redirect(sorteo_path('/admin'))


@app.route('/admin/dudoso/<int:participante_id>/aprobar', methods=['POST'])
@app.route('/sorteo/admin/dudoso/<int:participante_id>/aprobar', methods=['POST'])
def aprobar_dudoso(participante_id):
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    conn = get_db()
    c = conn.cursor()
    c.execute('''
        UPDATE sorteo_participantes
        SET estado = 'APROBADO',
            detalle_validacion = 'Aprobado manualmente desde la bandeja de dudosos.',
            consulta_at = CURRENT_TIMESTAMP
        WHERE id = %s AND estacion_id = %s AND conteo_id IS NULL
    ''', (participante_id, session['sorteo_estacion_id']))
    conn.commit()
    conn.close()
    return redirect(sorteo_path('/admin'))


@app.route('/admin/dudoso/<int:participante_id>/denegar', methods=['POST'])
@app.route('/sorteo/admin/dudoso/<int:participante_id>/denegar', methods=['POST'])
def denegar_dudoso(participante_id):
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    conn = get_db()
    c = conn.cursor()
    c.execute('''
        UPDATE sorteo_participantes
        SET estado = 'DENEGADO',
            chances = 0,
            detalle_validacion = 'Denegado manualmente desde la bandeja de dudosos.',
            consulta_at = CURRENT_TIMESTAMP
        WHERE id = %s AND estacion_id = %s AND conteo_id IS NULL
    ''', (participante_id, session['sorteo_estacion_id']))
    conn.commit()
    conn.close()
    return redirect(sorteo_path('/admin'))


@app.route('/admin/denegado/<int:participante_id>/aprobar', methods=['POST'])
@app.route('/sorteo/admin/denegado/<int:participante_id>/aprobar', methods=['POST'])
def aprobar_denegado(participante_id):
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute('''
        SELECT *
        FROM sorteo_participantes
        WHERE id = %s AND estacion_id = %s AND conteo_id IS NULL AND estado = 'DENEGADO'
    ''', (participante_id, session['sorteo_estacion_id']))
    row = c.fetchone()
    if row:
        chances = compute_chances(row.get('combustible'), bool(row.get('pago_app_ypf'))) if row.get('combustible') else 1
        c.execute('''
            UPDATE sorteo_participantes
            SET estado = 'APROBADO',
                chances = %s,
                detalle_validacion = 'Aprobado manualmente desde la bandeja de denegados.',
                consulta_at = CURRENT_TIMESTAMP
            WHERE id = %s AND estacion_id = %s AND conteo_id IS NULL
        ''', (chances, participante_id, session['sorteo_estacion_id']))
        conn.commit()
    conn.close()
    return redirect(sorteo_path('/admin'))


@app.route('/admin/reprocesar-facturas', methods=['POST'])
@app.route('/sorteo/admin/reprocesar-facturas', methods=['POST'])
def reprocesar_facturas():
    if not admin_required():
        return jsonify({'ok': False, 'error': 'Sesion expirada.'}), 401

    payload = request.get_json(silent=True) or {}
    items = payload.get('items') or []
    if not items:
        return jsonify({'ok': False, 'error': 'No se recibieron filas para reprocesar.'}), 400

    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    actualizados = []
    errores = []
    try:
        for item in items:
            participante_id = item.get('id')
            numero_factura = item.get('numero_factura')
            ticket_fecha = item.get('ticket_fecha')
            try:
                canonical, normalized_fecha = requeue_participante_with_invoice(
                    c,
                    session['sorteo_estacion_id'],
                    int(participante_id),
                    numero_factura,
                    ticket_fecha=ticket_fecha,
                )
                actualizados.append({'id': int(participante_id), 'numero_factura': canonical, 'ticket_fecha': normalized_fecha})
            except Exception as exc:
                errores.append({'id': participante_id, 'error': str(exc)})
        conn.commit()
    except Exception as exc:
        conn.rollback()
        conn.close()
        return jsonify({'ok': False, 'error': str(exc)}), 500

    conn.close()
    return jsonify({'ok': True, 'actualizados': actualizados, 'errores': errores})


@app.route('/admin/participante/<int:participante_id>/eliminar', methods=['POST'])
@app.route('/sorteo/admin/participante/<int:participante_id>/eliminar', methods=['POST'])
def eliminar_participante(participante_id):
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))

    conn = get_db()
    c = conn.cursor()
    c.execute('''
        DELETE FROM sorteo_participantes
        WHERE id = %s
          AND estacion_id = %s
          AND conteo_id IS NULL
          AND estado IN ('APROBADO', 'DENEGADO', 'DUDOSO')
    ''', (participante_id, session['sorteo_estacion_id']))
    conn.commit()
    conn.close()
    return redirect(sorteo_path('/admin'))


@app.route('/admin/participante/<int:participante_id>/detalle-ticket')
@app.route('/sorteo/admin/participante/<int:participante_id>/detalle-ticket')
def detalle_ticket_participante(participante_id):
    if not admin_required():
        return jsonify({'ok': False, 'error': 'Sesion expirada.'}), 401

    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute('''
        SELECT id, estacion_id, ticket_fecha, numero_factura, factura_real
        FROM sorteo_participantes
        WHERE id = %s
          AND estacion_id = %s
    ''', (participante_id, session['sorteo_estacion_id']))
    row = c.fetchone()
    conn.close()

    if not row:
        return jsonify({'ok': False, 'error': 'Participante no encontrado.'}), 404

    numero_factura = row.get('factura_real') or row.get('numero_factura')
    detalle_rows, error = consultar_detalle_ticket(
        row['estacion_id'],
        row['ticket_fecha'],
        numero_factura,
    )
    if error:
        return jsonify({'ok': False, 'error': error}), 400

    return jsonify({
        'ok': True,
        'fecha': str(row['ticket_fecha']),
        'factura': numero_factura,
        'items': format_ticket_detail_rows(detalle_rows),
    })


@app.route('/admin/exportar-excel')
@app.route('/sorteo/admin/exportar-excel')
def exportar_excel():
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    eid = session['sorteo_estacion_id']
    incluir_archivados = request.args.get('todo') == '1'
    estado = (request.args.get('estado') or '').upper().strip()
    ponderado = request.args.get('ponderado') == '1'

    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    rows = query_participantes(c, eid, archived=incluir_archivados)
    conn.close()

    if estado:
        rows = [row for row in rows if row['estado'] == estado]

    export_rows = []
    for row in rows:
        repeats = max(int(row.get('chances') or 1), 1) if ponderado and row['estado'] == 'APROBADO' else 1
        for idx in range(repeats):
            export_rows.append((row, idx + 1))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sorteo Electrodomesticos'
    ws.append(EXPORT_HEADERS)
    for row, repetition in export_rows:
        ws.append(build_export_values(row, repetition if ponderado and row['estado'] == 'APROBADO' else ''))
    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    suffix = estado.lower() if estado else 'general'
    if incluir_archivados:
        suffix += '_archivado'
    if ponderado:
        suffix += '_ponderado'
    return Response(
        salida.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment;filename=sorteo_{suffix}.xlsx'},
    )


@app.route('/admin/exportar-archivado/<int:conteo_id>')
@app.route('/sorteo/admin/exportar-archivado/<int:conteo_id>')
def exportar_archivado(conteo_id):
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))

    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute('''
        SELECT *
        FROM sorteo_conteos
        WHERE id = %s
          AND estacion_id = %s
    ''', (conteo_id, session['sorteo_estacion_id']))
    conteo = c.fetchone()
    conn.close()

    if not conteo:
        return 'Archivado no encontrado.', 404

    snapshot = conteo.get('snapshot_json') or '[]'
    try:
        rows = json.loads(snapshot)
    except Exception:
        rows = []
    ranking_snapshot = conteo.get('ranking_json') or '[]'
    try:
        ranking_rows = json.loads(ranking_snapshot)
    except Exception:
        ranking_rows = []

    grouped = {
        'APROBADO': [row for row in rows if (row.get('estado') or '').upper() == 'APROBADO'],
        'DENEGADO': [row for row in rows if (row.get('estado') or '').upper() == 'DENEGADO'],
        'PENDIENTE': [row for row in rows if (row.get('estado') or '').upper() == 'PENDIENTE'],
        'DUDOSO': [row for row in rows if (row.get('estado') or '').upper() == 'DUDOSO'],
    }
    if not ranking_rows:
        ranking_rows = build_seller_ranking(grouped['APROBADO'])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    append_archive_summary_sheet(wb, conteo)
    append_archive_ranking_sheet(wb, ranking_rows)
    append_export_sheet(wb, 'Aprobados', grouped['APROBADO'], ponderado=False)
    append_export_sheet(wb, 'Denegados', grouped['DENEGADO'], ponderado=False)
    append_export_sheet(wb, 'Pendientes', grouped['PENDIENTE'], ponderado=False)
    append_export_sheet(wb, 'Dudosos', grouped['DUDOSO'], ponderado=False)

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    fecha_archivo = str(conteo.get('iniciado_en') or '')[:10] or 'archivado'
    return Response(
        salida.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment;filename=sorteo_archivado_{conteo_id}_{fecha_archivo}.xlsx'},
    )


@app.route('/admin/exportar-ranking')
@app.route('/sorteo/admin/exportar-ranking')
def exportar_ranking():
    if not admin_required():
        return redirect(sorteo_path('/admin/login'))
    eid = session['sorteo_estacion_id']
    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    rows = query_participantes(c, eid, archived=False)
    conn.close()
    ranking = build_seller_ranking([row for row in rows if row['estado'] == 'APROBADO'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ranking Vendedores'
    ws.append(['Puesto', 'Vendedor', 'Facturas aprobadas'])
    for index, row in enumerate(ranking, start=1):
        ws.append([index, row['vendedor'], row['facturas_aprobadas']])

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return Response(
        salida.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=sorteo_ranking_vendedores.xlsx'},
    )


@app.route('/<int:estacion_id>', methods=['GET', 'POST'])
@app.route('/sorteo/<int:estacion_id>', methods=['GET', 'POST'])
@app.route('/participar/<int:estacion_id>', methods=['GET', 'POST'])
def cliente_sorteo(estacion_id):
    ensure_config(estacion_id)
    conn = get_db()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute('SELECT nombre FROM estaciones WHERE id = %s', (estacion_id,))
    estacion = c.fetchone()
    c.execute('SELECT * FROM sorteo_config WHERE estacion_id = %s', (estacion_id,))
    config = c.fetchone()
    mensaje = None
    error = None

    if not estacion:
        conn.close()
        return 'Estacion no encontrada', 404

    if request.method == 'POST':
        if config and config['detenido']:
            error = 'El sorteo se encuentra cerrado.'
        else:
            try:
                if request.form.get('acepta_promociones') != '1':
                    raise ValueError('Necesitas aceptar el contacto por promociones futuras para participar.')
                ticket_fecha = parse_ticket_date(request.form['fecha_ticket'])
                promo_desde, promo_hasta = get_active_promo_bounds(config)
                if ticket_fecha < promo_desde or ticket_fecha > promo_hasta:
                    raise ValueError('La fecha del ticket esta fuera de la promocion activa.')
                factura = parse_invoice_from_form(request.form)
                validate_allowed_invoice_branch(estacion, factura)
                device_token = (request.form.get('device_token') or '').strip()
                user_agent = request.headers.get('User-Agent', '')
                ip_registro = get_client_ip()
                sospecha = count_device_registrations(c, estacion_id, config, device_token) >= 2
                estado_inicial = 'PENDIENTE'
                detalle = 'Esperando validacion.' if not sospecha else 'Esperando validacion. El dispositivo sera revisado si la factura matchea.'
                c.execute('''
                    INSERT INTO sorteo_participantes
                    (estacion_id, ticket_fecha, ticket_hora, numero_factura, nombre, apellido, dni, telefono, email, acepta_promociones, estado, device_token, ip_registro, user_agent, sospecha_dispositivo, detalle_validacion)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    estacion_id,
                    ticket_fecha,
                    datetime.combine(ticket_fecha, datetime.min.time()),
                    factura['canonical'],
                    request.form['nombre'].strip(),
                    request.form['apellido'].strip(),
                    request.form.get('dni', '').strip(),
                    request.form['telefono'].strip(),
                    request.form['email'].strip().lower(),
                    True,
                    estado_inicial,
                    device_token,
                    ip_registro,
                    user_agent,
                    sospecha,
                    detalle,
                ))
                conn.commit()
                mensaje = 'Tu cupon quedo registrado. Si la factura coincide, pasara a aprobado; si detectamos algo raro con el dispositivo, quedara en revision.'
            except psycopg2.errors.UniqueViolation:
                conn.rollback()
                error = 'Ese ticket ya fue registrado para esta estacion.'
            except ValueError as exc:
                error = str(exc)
            except Exception as exc:
                conn.rollback()
                error = f'No se pudo registrar la participacion: {exc}'
    conn.close()
    return render_template(
        'sorteo_cliente.html',
        estacion=estacion,
        mensaje=mensaje,
        error=error,
        detenido=config and config['detenido'],
        allowed_branches=get_allowed_invoice_branches(estacion),
        factura_help_image='/static/img/factura-ayuda-sorteo.png',
    )


init_db()

if os.environ.get('SORTEO_DISABLE_SCHEDULER') != '1':
    threading.Thread(target=scheduler_loop, daemon=True).start()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('SORTEO_PORT', '5055')))
